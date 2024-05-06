from PyQt5.QtWidgets import QMainWindow ,QApplication,QTableWidgetItem,QMessageBox
from PyQt5.uic import loadUi
from PyQt5 import QtWidgets,QtGui
import openpyxl
import sys
import os
import re
from delete import remove


class MainUi(QMainWindow):
    def __init__(self):

        super(MainUi,self).__init__()
        loadUi("ui\main.ui",self)

        self.actionAjouter_Etudiant.triggered.connect(self.ajouter_etud_screen)
        self.actionSupprimer_Etudiant.triggered.connect(self.supp_etud)
        self.actionModifier_Etudiant.triggered.connect(self.Modifier_etud)
        self.actionContenue_du_dictionnaire_Etudiant.triggered.connect(self.Afficher_etud)
        self.actionRecherche.triggered.connect(self.rech)
        
        
        
    def ajouter_etud_screen(self):
        Ajouter_etudient = Ajouter_etud()
        widget.addWidget(Ajouter_etudient)
        widget.setCurrentIndex(widget.currentIndex()+1)
        self.statusbar.showMessage("ajouter etud clicked")  
    def supp_etud(self):
        Supprimer_etudiant = Supprimer_etud()
        widget.addWidget(Supprimer_etudiant)
        widget.setCurrentIndex(widget.currentIndex()+1)
        
    

    def Modifier_etud(self):
        Modifier_etudiante = Modifier_etudiant()
        widget.addWidget(Modifier_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def rech(self):
        rech = recherche_etud()
        widget.addWidget(rech)
        widget.setCurrentIndex(widget.currentIndex()+1)    
    def Afficher_etud(self):    
        afficher_etudiante = Afficher_etudiants()
        widget.addWidget(afficher_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)



class Ajouter_etud(QMainWindow):
    def __init__(self):

        super(Ajouter_etud,self).__init__()
        loadUi("ui\Ajouter_etud.ui",self) 
        self.actionSupprimer_Etudiant.triggered.connect(self.supp_etud)  
        self.actionModifier_Etudiant.triggered.connect(self.Modifier_etud) 
        self.actionContenue_du_dictionnaire_Etudiant.triggered.connect(self.Afficher_etud)
        self.pushButton.clicked.connect(self.send)   
        self.actionRecherche.triggered.connect(self.rech)
        
       
   
    def supp_etud(self):
        Supprimer_etudiant = Supprimer_etud()
        widget.addWidget(Supprimer_etudiant)
        widget.setCurrentIndex(widget.currentIndex()+1)
        self.statusbar.showMessage("suppresion eleve") 
    def rech(self):
        Rech = recherche_etud()
        widget.addWidget(Rech)
        widget.setCurrentIndex(widget.currentIndex()+1)    
    def clear(self):
        path="data\data.xlsx"
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        for row in sheet:
            remove(sheet)
        workbook.save(path)       
    def Afficher_etud(self):    
        afficher_etudiante = Afficher_etudiants()
        widget.addWidget(afficher_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)     

    def Modifier_etud(self):
        Modifier_etudiante = Modifier_etudiant()
        widget.addWidget(Modifier_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1) 
    def popup_suc(self):
        msg=QMessageBox()    
        msg.setWindowTitle("Success")
        msg.setText("eleve enregistree avec succes !")
        msg.setIcon(QMessageBox.Information)
       

        x=msg.exec_()
    def popup_erre(self):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Critical)
        msg.setText("errur d information !")
        msg.setWindowTitle("ERREUR")
        msg.setDetailedText("-Numero inscription doit contenir 8 chiffres \n -phone doit etre formee par 8 chiffres\n -numero de rue doit etre numerique \n -email doit contenir @ et .  ")
        
        

        x=msg.exec_()    
    def send(self):
        tito=True
        n_inscri=self.lineEdit.text()
        nom=self.lineEdit_2.text()
        prenom=self.lineEdit_3.text()
        mail=self.lineEdit_5.text()
        phone=self.lineEdit_6.text()
        n_rue=self.lineEdit_7.text()
        nom_rue=self.lineEdit_8.text()
        ville=self.lineEdit_9.text()
        date=self.dateEdit.text()
        section=self.comboBox.currentText()
        niveau_etud=self.comboBox_2.currentText()
        
        match = re.match('^[_a-z0-9-]+(\.[_a-z0-9-]+)*@[a-z0-9-]+(\.[a-z0-9-]+)*(\.[a-z]{2,4})$', mail)
        if(len(n_inscri)!=8 or n_inscri.isnumeric()==False):
            tito=False
           
        if(len(phone)!=8 or phone.isnumeric()==False):
            tito=False  
          
        if(nom.isalpha()==False or prenom.isalpha()==False or nom_rue.isalpha()==False or ville.isalpha()==False):
            tito=False        
        if (match==None):
            tito=False
        if(n_rue.isnumeric()==False):
            tito=False 
        d=date[6:]    
        if(2023<int(d)<1800):
            tito=False        
        print("---------------")
        print("numero d inscrit: ",n_inscri,"|nom : ",nom,"|prenom :",prenom,"Date :",date,"mail :",mail,
              "phone :",phone,"adresse :",n_rue+" "+nom_rue+" "+ville,"|niveau etude  :",niveau_etud+section)
        print("---------------")
        self.lineEdit.setText('')
        self.lineEdit_2.setText('')
        self.lineEdit_3.setText('')
        self.lineEdit_5.setText('')
        self.lineEdit_6.setText('')
        self.lineEdit_7.setText('')
        self.lineEdit_8.setText('')
        self.lineEdit_9.setText('')
        path="data\data.xlsx"
        if not os.path.exists(path):
            workbook=openpyxl.Workbook()
            sheet=workbook.active
            heading=["N° inscri","Nom","Prenom","Date Naissance","Mail","Telephone","Adresse","Niveau d Etude","Section"]
            sheet.append(heading)
            workbook.save(path)
        if(tito==True):    
            workbook=openpyxl.load_workbook(path)
            sheet=workbook.active
            sheet.append([n_inscri,nom,prenom,date,mail,phone,n_rue+" "+nom_rue+" "+ville,niveau_etud,section])   
            self.popup_suc() 
            self.clear()
            workbook.save(path)  
            
        else:
            print("erreur") 
            self.popup_erre()  

                                            # SUPPRIMER ETUDIANT #   


                  
class Supprimer_etud(QMainWindow):
    def __init__(self):

        super(Supprimer_etud,self).__init__()
        loadUi("ui\Supprimer_etudiant.ui",self)  
        self.actionAjouter_Etudiant.triggered.connect(self.ajouter_etud_screen) 
        self.actionModifier_Etudiant.triggered.connect(self.Modifier_etud)
        self.actionContenue_du_dictionnaire_Etudiant.triggered.connect(self.Afficher_etud)
        self.pushButton.clicked.connect(self.supprimer_insc)
        self.pushButton_2.clicked.connect(self.supprimer_sec)
        self.pushButton_3.clicked.connect(self.supprimer_niv)
        self.pushButton_4.clicked.connect(self.supprimer_niv_sec)
        self.actionRecherche_par_numero_inscription.triggered.connect(self.rech)
    def ajouter_etud_screen(self):
        Ajouter_etudient = Ajouter_etud()
        widget.addWidget(Ajouter_etudient)
        widget.setCurrentIndex(widget.currentIndex()+1)
        self.statusbar.showMessage("ajouter etud clicked")    
    def Modifier_etud(self):
        Modifier_etudiante = Modifier_etudiant()
        widget.addWidget(Modifier_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1) 
    def rech(self):
        Rech = recherche_etud()
        widget.addWidget(Rech)
        widget.setCurrentIndex(widget.currentIndex()+1)     
    def Afficher_etud(self):    
        afficher_etudiante = Afficher_etudiants()
        widget.addWidget(afficher_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)    
    def popup_ins(self):
        msg=QMessageBox()    
        msg.setWindowTitle("Success")
        msg.setText(f"eleve {self.lineEdit.text()} a ete supprimee avec succes !")
        msg.setIcon(QMessageBox.Information)
        x=msg.exec_()
    def popup_sec(self):
        msg=QMessageBox()    
        msg.setWindowTitle("Success")
        msg.setText(f"Section {self.comboBox.currentText()} a ete supprimee avec succes !")
        msg.setIcon(QMessageBox.Information)
        x=msg.exec_()
           
    def popup_niv(self):
        msg=QMessageBox()    
        msg.setWindowTitle("Success")
        msg.setText(f" Niveau {self.comboBox_2.currentText()} annee a ete supprimee avec succes !")
        msg.setIcon(QMessageBox.Information)
        x=msg.exec_()
    def popup_niv_sec(self):
        msg=QMessageBox()    
        msg.setWindowTitle("Success")
        msg.setText(f" {self.comboBox_3.currentText()} {self.comboBox_4.currentText()}  a ete supprimee avec succes !")
        msg.setIcon(QMessageBox.Information)
        x=msg.exec_()  

    def supprimer_insc(self):
        path="data\data.xlsx"
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        for row in range(2,sheet.max_row):
            if(sheet["A"+str(row)].value ==self.lineEdit.text()):
                sheet.delete_rows(row)
                workbook.save(path) 
                print("deleted")
        self.lineEdit.setText('')        
        self.popup_ins()
    def supprimer_sec(self):
        path="data\data.xlsx"
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        for row in range(2,sheet.max_row):
            while(sheet["I"+str(row)].value==self.comboBox.currentText()):
                sheet.delete_rows(row)
                print("row deleted") 
        self.popup_sec()        
        workbook.save(path)               
    def supprimer_niv(self):
        path="data\data.xlsx"
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
         
        for row in range(2,sheet.max_row):
            while(sheet["H"+str(row)].value ==self.comboBox_2.currentText()):
                print("row deleted") 
                sheet.delete_rows(row)
        self.popup_niv()              
        workbook.save(path)   
    def supprimer_niv_sec(self):
        path="data\data.xlsx"
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
         
        for row in range(2,sheet.max_row):
            while(sheet["H"+str(row)].value ==self.comboBox_3.currentText() and sheet["I"+str(row)].value==self.comboBox_4.currentText()):
                print("row deleted") 
                sheet.delete_rows(row)  
        self.popup_niv_sec()            
        workbook.save(path)               
            

       


                                            # MODIFIER ETUDIANT #
class Modifier_etudiant(QMainWindow):
    def __init__(self):

        super(Modifier_etudiant,self).__init__()
        loadUi("ui\Modifier_etud.ui",self)   
        self.actionAjouter_Etudiant.triggered.connect(self.ajouter_etud_screen)
        self.actionSupprimer_Etudiant.triggered.connect(self.supp_etud) 
        self.actionContenue_du_dictionnaire_Etudiant.triggered.connect(self.Afficher_etud)
        self.pushButton.clicked.connect(self.modifier_tel)
        self.pushButton_2.clicked.connect(self.modifier_mail)
        self.pushButton_3.clicked.connect(self.modifier_adress)
       
    def ajouter_etud_screen(self):
        Ajouter_etudient = Ajouter_etud()
        widget.addWidget(Ajouter_etudient)
        widget.setCurrentIndex(widget.currentIndex()+1)
        self.statusbar.showMessage("ajouter etud clicked")   
    def Afficher_etud(self):    
        afficher_etudiante = Afficher_etudiants()
        widget.addWidget(afficher_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)    
    def supp_etud(self):
        Supprimer_etudiant = Supprimer_etud()
        widget.addWidget(Supprimer_etudiant)
        widget.setCurrentIndex(widget.currentIndex()+1)
        self.statusbar.showMessage("suppresion eleve")
    def rech(self):
        rech = recherche_etud()
        widget.addWidget(rech)
        widget.setCurrentIndex(widget.currentIndex()+1)     
    def popup_md_tel(self):
        msg=QMessageBox()    
        msg.setWindowTitle("Success")
        msg.setText(f" Numero Telephone de  {self.lineEdit_11.text()}  a ete modifiee avec succes !")
        msg.setIcon(QMessageBox.Information)  
        x=msg.exec_()   
    def modifier_tel(self):
        path="data\data.xlsx"
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
         
        for row in range(2,sheet.max_row):
            if(sheet["A"+str(row)].value==self.lineEdit_11.text()):
                sheet["F"+str(row)].value=self.lineEdit.text()
                
                print("done")
        self.lineEdit_11.setText('')  
        self.lineEdit.setText('')         
        self.popup_md_tel() 
        workbook.save(path) 
                                            #modifier email #        
    def popup_md_mail(self):
        msg=QMessageBox()    
        msg.setWindowTitle("Success")
        msg.setText(f" Mail de  {self.lineEdit_3.text()}  a ete modifiee avec succes !")
        msg.setIcon(QMessageBox.Information)  
        x=msg.exec_()     
    def modifier_mail(self):
        path="data\data.xlsx"
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
         
        for row in range(2,sheet.max_row):
            if(sheet["A"+str(row)].value==self.lineEdit_3.text()):
                sheet["E"+str(row)].value=self.lineEdit_2.text()
                print("done")
        self.popup_md_mail() 
        self.lineEdit_3.setText('') 
        self.lineEdit_2.setText('') 
        workbook.save(path)  
                                        #Adresse modification#
    def popup_adress(self):
        msg=QMessageBox()    
        msg.setWindowTitle("Success")
        msg.setText(f" Adresse de  {self.lineEdit_10.text()}  a ete modifiee avec succes !")
        msg.setIcon(QMessageBox.Information)  
        x=msg.exec_()     
    def modifier_adress(self):
        path="data\data.xlsx"
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
         
        for row in range(2,sheet.max_row):
            if(sheet["A"+str(row)].value==self.lineEdit_10.text()):
                sheet["G"+str(row)].value=self.lineEdit_7.text()+self.lineEdit_8.text()+self.lineEdit_9.text()
                
                print("done")
        self.popup_adress() 
        self.lineEdit_10.setText('') 
        self.lineEdit_7.setText('') 
        self.lineEdit_8.setText('') 
        self.lineEdit_9.setText('') 
        workbook.save(path)  
               
class recherche_etud(QMainWindow):
    def __init__(self):
        super(recherche_etud,self).__init__()
        loadUi("ui\Recherche.ui",self)
        self.actionAjouter_Etudiant.triggered.connect(self.ajouter_etud_screen)
        self.actionSupprimer_Etudiant.triggered.connect(self.supp_etud) 
        self.actionModifier_Etudiant.triggered.connect(self.Modifier_etud)
        self.actionContenue_du_dictionnaire_Etudiant.triggered.connect(self.Afficher_etud)
        self.pushButton_2.clicked.connect(self.load_inscri)
        self.pushButton_3.clicked.connect(self.load_Nom)
        self.pushButton_4.clicked.connect(self.load_Prenom)
        self.pushButton_5.clicked.connect(self.load_Section)
        self.pushButton_6.clicked.connect(self.load_Niveau)
        self.pushButton.clicked.connect(self.load_Tel)
        self.pushButton_7.clicked.connect(self.load_Mail)
        self.pushButton_8.clicked.connect(self.load_Adress)
    def ajouter_etud_screen(self):
        Ajouter_etudient = Ajouter_etud()
        widget.addWidget(Ajouter_etudient)
        widget.setCurrentIndex(widget.currentIndex()+1)
        self.statusbar.showMessage("ajouter etud clicked")  
    def supp_etud(self):
        Supprimer_etudiant = Supprimer_etud()
        widget.addWidget(Supprimer_etudiant)
        widget.setCurrentIndex(widget.currentIndex()+1)
        self.statusbar.showMessage("suppresion eleve") 
    def load_inscri(self):
        n_inscri=self.lineEdit_2.text()
        self.tableWidget.setColumnCount(9)
        self.tableWidget.setRowCount(200)
        self.tableWidget.setHorizontalHeaderLabels(("N° Inscri ","Nom","Prenom","Date Naissance","Mail","Telephone","Adresse","Niveau d etude","Section"))
        path="data\data.xlsx"    
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        list_value=list(sheet.values)
        row_ind=0
        for vt in list_value[1:]:
            col_ind=0
            for v in vt:
                self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(""))
                col_ind+=1                         
            row_ind+=1 
        row_ind=0
        for row in sheet.iter_rows(min_row=2, max_col=9, max_row=99, values_only=True):
            for cell in row:
                if cell==n_inscri:
                    col_ind=0
                    for v in row:
                        self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(str(v)))
                        col_ind+=1                         
                    row_ind+=1  
    def load_Nom(self):
        nom=self.lineEdit_3.text()
        self.tableWidget.setColumnCount(9)
        self.tableWidget.setRowCount(200)
        self.tableWidget.setHorizontalHeaderLabels(("N° Inscri ","Nom","Prenom","Date Naissance","Mail","Telephone","Adresse","Niveau d etude","Section"))
        path="data\data.xlsx"    
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        list_value=list(sheet.values)
        row_ind=0
        for vt in list_value[1:]:
            col_ind=0
            for v in vt:
                self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(""))
                col_ind+=1                         
            row_ind+=1         
        row_ind=0
        for row in sheet.iter_rows(min_row=2, max_col=9, max_row=99, values_only=True):
            for cell in row:
                if cell==nom:
                    col_ind=0
                    for v in row:
                        self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(str(v)))
                        col_ind+=1     

                    row_ind+=1 
                  
                                   
    def load_Prenom(self):
        prenom=self.lineEdit_4.text()
        self.tableWidget.setColumnCount(9)
        self.tableWidget.setRowCount(200)
        self.tableWidget.setHorizontalHeaderLabels(("N° Inscri ","Nom","Prenom","Date Naissance","Mail","Telephone","Adresse","Niveau d etude","Section"))
        path="data\data.xlsx"    
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        list_value=list(sheet.values)
        row_ind=0
        for vt in list_value[1:]:
            col_ind=0
            for v in vt:
                self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(""))
                col_ind+=1                         
            row_ind+=1 
        row_ind=0
        for row in sheet.iter_rows(min_row=2, max_col=9, max_row=99, values_only=True):
            for cell in row:
                if cell==prenom:
                    col_ind=0
                    for v in row:
                        self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(str(v)))
                        col_ind+=1                         
                    row_ind+=1                               
    def load_Section(self):
        section=self.lineEdit_5.text()
        self.tableWidget.setColumnCount(9)
        self.tableWidget.setRowCount(200)
        self.tableWidget.setHorizontalHeaderLabels(("N° Inscri ","Nom","Prenom","Date Naissance","Mail","Telephone","Adresse","Niveau d etude","Section"))
        path="data\data.xlsx"    
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        list_value=list(sheet.values)
        row_ind=0
        for vt in list_value[1:]:
            col_ind=0
            for v in vt:
                self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(""))
                col_ind+=1                         
            row_ind+=1 
        row_ind=0
        for row in sheet.iter_rows(min_row=2, max_col=9, max_row=99, values_only=True):
            for cell in row:
                if cell==section:
                    col_ind=0
                    for v in row:
                        self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(str(v)))
                        col_ind+=1                         
                    row_ind+=1    
    def load_Niveau(self):
        niv=self.lineEdit_6.text()
        self.tableWidget.setColumnCount(9)
        self.tableWidget.setRowCount(200)
        self.tableWidget.setHorizontalHeaderLabels(("N° Inscri ","Nom","Prenom","Date Naissance","Mail","Telephone","Adresse","Niveau d etude","Section"))
        path="data\data.xlsx"    
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        list_value=list(sheet.values)
        row_ind=0
        for vt in list_value[1:]:
            col_ind=0
            for v in vt:
                self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(""))
                col_ind+=1                         
            row_ind+=1 
        row_ind=0
        for row in sheet.iter_rows(min_row=2, max_col=9, max_row=99, values_only=True):
            for cell in row:
                if cell==niv:
                    col_ind=0
                    for v in row:
                        self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(str(v)))
                        col_ind+=1                         
                    row_ind+=1                           
    def load_Tel(self):
        tel=self.lineEdit.text()
        self.tableWidget.setColumnCount(9)
        self.tableWidget.setRowCount(200)
        self.tableWidget.setHorizontalHeaderLabels(("N° Inscri ","Nom","Prenom","Date Naissance","Mail","Telephone","Adresse","Niveau d etude","Section"))
        path="data\data.xlsx"    
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        list_value=list(sheet.values)
        row_ind=0
        for vt in list_value[1:]:
            col_ind=0
            for v in vt:
                self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(""))
                col_ind+=1                         
            row_ind+=1 
        row_ind=0
        for row in sheet.iter_rows(min_row=2, max_col=9, max_row=99, values_only=True):
            for cell in row:
                if cell==tel:
                    col_ind=0
                    for v in row:
                        self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(str(v)))
                        col_ind+=1                         
                    row_ind+=1    
    def load_Mail(self):
        mail=self.lineEdit_7.text()
        self.tableWidget.setColumnCount(9)
        self.tableWidget.setRowCount(200)
        self.tableWidget.setHorizontalHeaderLabels(("N° Inscri ","Nom","Prenom","Date Naissance","Mail","Telephone","Adresse","Niveau d etude","Section"))
        path="data\data.xlsx"    
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        list_value=list(sheet.values)
        row_ind=0
        for vt in list_value[1:]:
            col_ind=0
            for v in vt:
                self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(""))
                col_ind+=1                         
            row_ind+=1 
        row_ind=0
        for row in sheet.iter_rows(min_row=2, max_col=9, max_row=99, values_only=True):
            for cell in row:
                if cell==mail:
                    col_ind=0
                    for v in row:
                        self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(str(v)))
                        col_ind+=1                         
                    row_ind+=1   
    def load_Adress(self):
        adr=self.lineEdit_8.text()
        self.tableWidget.setColumnCount(9)
        self.tableWidget.setRowCount(200)
        self.tableWidget.setHorizontalHeaderLabels(("N° Inscri ","Nom","Prenom","Date Naissance","Mail","Telephone","Adresse","Niveau d etude","Section"))
        path="data\data.xlsx"    
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        list_value=list(sheet.values)
        row_ind=0
        for vt in list_value[1:]:
            col_ind=0
            for v in vt:
                self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(""))
                col_ind+=1                         
            row_ind+=1 
        row_ind=0
        for row in sheet.iter_rows(min_row=2, max_col=9, max_row=99, values_only=True):
            for cell in row:
                if cell==adr:
                    col_ind=0
                    for v in row:
                        self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(str(v)))
                        col_ind+=1                         
                    row_ind+=1                                      
    def Modifier_etud(self):
        Modifier_etudiante = Modifier_etudiant()
        widget.addWidget(Modifier_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def Afficher_etud(self):    
        afficher_etudiante = Afficher_etudiants()
        widget.addWidget(afficher_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)   

class Afficher_etudiants(QMainWindow):
    def __init__(self):
        super(Afficher_etudiants,self).__init__()
        loadUi("ui\Afficher_etud.ui",self)
        self.actionAjouter_Etudiant.triggered.connect(self.ajouter_etud_screen)
        self.actionSupprimer_Etudiant.triggered.connect(self.supp_etud) 
        self.actionModifier_Etudiant.triggered.connect(self.Modifier_etud)
        self.actionRecherche_par_numero_inscription.triggered.connect(self.rech)
        
        self.load_data()

    def ajouter_etud_screen(self):
        Ajouter_etudient = Ajouter_etud()
        widget.addWidget(Ajouter_etudient)
        widget.setCurrentIndex(widget.currentIndex()+1)
        self.statusbar.showMessage("ajouter etud clicked")  
    def rech(self):
        rech = recherche_etud()
        widget.addWidget(rech)
        widget.setCurrentIndex(widget.currentIndex()+1)        
    def rech(self):
        rech = recherche_etud()
        widget.addWidget(rech)
        widget.setCurrentIndex(widget.currentIndex()+1)      
    def supp_etud(self):
        Supprimer_etudiant = Supprimer_etud()
        widget.addWidget(Supprimer_etudiant)
        widget.setCurrentIndex(widget.currentIndex()+1)
        self.statusbar.showMessage("suppresion eleve")    
    def Modifier_etud(self):
        Modifier_etudiante = Modifier_etudiant()
        widget.addWidget(Modifier_etudiante)
        widget.setCurrentIndex(widget.currentIndex()+1)   
   
    
    def load_data(self):
        self.tableWidget.setColumnCount(9)
        self.tableWidget.setRowCount(200)
        self.tableWidget.setHorizontalHeaderLabels(("N° Inscri ","Nom","Prenom","Date Naissance","Mail","Telephone","Adresse","Niveau d etude","Section"))
        path="data\data.xlsx"    
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.active
        list_value=list(sheet.values)
        row_ind=0
        for vt in list_value[1:]:
            col_ind=0
            for v in vt:
                print(v)

                self.tableWidget.setItem(row_ind,col_ind,QTableWidgetItem(str(v)))
                col_ind+=1                         
            row_ind+=1                            
           


if __name__=='__main__':
     app=QApplication(sys.argv)
     ui=MainUi()
     MainWindow=MainUi()
     widget=QtWidgets.QStackedWidget()
     widget.addWidget(MainWindow)
     widget.setWindowTitle('ISIMM Library')
     widget.setWindowIcon(QtGui.QIcon('ui\images\lg2.png'))
     widget.setFixedHeight(600)
     widget.setFixedWidth(1250)
     widget.show()
     app.exec_()            